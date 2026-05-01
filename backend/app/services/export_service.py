from __future__ import annotations

from io import BytesIO, StringIO

import ezdxf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from sqlalchemy.orm import Session

from app.services.calculation_service import get_results

# ─── yardımcı ────────────────────────────────────────────────────────────────

_PHASE_COLORS: dict[str, str] = {
    "L1": "FFB300",   # amber
    "L2": "E53935",   # red
    "L3": "1E88E5",   # blue
    "N":  "43A047",   # green
    "PE": "6D4C41",   # brown
}

_HEADER_FILL = PatternFill("solid", fgColor="1E293B")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _header_row(sheet, values: list[str]) -> None:
    sheet.append(values)
    for cell in sheet[sheet.max_row]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


# ─── CSV ─────────────────────────────────────────────────────────────────────

def build_csv(db: Session, project_id: int) -> bytes:
    results = get_results(db, project_id)
    buf = StringIO()
    buf.write(
        "part_no,busbar_type,phase,connected_device,material,"
        "width_mm,thickness_mm,cut_length_mm,hole_count,bend_count\n"
    )
    for b in results.busbars:
        buf.write(
            f"{b.part_no},{b.busbar_type},{b.phase},"
            f"{b.connected_device_label or ''},"
            f"{b.material},{b.width_mm},{b.thickness_mm},"
            f"{b.cut_length_mm},{len(b.holes)},{len(b.bends)}\n"
        )
    return buf.getvalue().encode("utf-8")


# ─── Excel ───────────────────────────────────────────────────────────────────

def build_excel(db: Session, project_id: int) -> bytes:
    results = get_results(db, project_id)
    wb = Workbook()

    # ── Özet ─────────────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Ozet"
    _header_row(ws_sum, ["Metrik", "Deger"])
    ws_sum.append(["Ana Bakir Sayisi",   results.summary.main_busbar_count])
    ws_sum.append(["Tali Bakir Sayisi",  results.summary.branch_busbar_count])
    ws_sum.append(["Toplam Kesim Boyu (mm)", float(results.summary.total_cut_length_mm)])
    ws_sum.append(["Toplam Delik",       results.summary.total_hole_count])
    ws_sum.append(["Toplam Bukum",       results.summary.total_bend_count])
    ws_sum.append(["Toplam Agirlik (kg)", float(results.summary.total_weight_kg)])
    if results.warnings:
        ws_sum.append([])
        ws_sum.append(["Geometri Uyarilari", ""])
        for w in results.warnings:
            ws_sum.append(["", w])
    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 18

    # ── Bakırlar ─────────────────────────────────────────────────────────────
    ws_bar = wb.create_sheet("Bakirlar")
    _header_row(ws_bar, [
        "Parca No", "Tip", "Faz", "Bagli Cihaz", "Malzeme",
        "Genislik (mm)", "Kalinlik (mm)", "Kesim Boyu (mm)", "Delik", "Bukum",
    ])
    for b in results.busbars:
        ws_bar.append([
            b.part_no, b.busbar_type, b.phase,
            b.connected_device_label or "—", b.material,
            float(b.width_mm), float(b.thickness_mm), float(b.cut_length_mm),
            len(b.holes), len(b.bends),
        ])
    for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws_bar.column_dimensions[col].width = 16

    # ── Delikler ─────────────────────────────────────────────────────────────
    ws_hole = wb.create_sheet("Delikler")
    _header_row(ws_hole, [
        "Parca No", "Tip", "Faz", "Delik No",
        "X (mm)", "Y (mm)", "Cap (mm)", "Slot G (mm)", "Slot U (mm)",
        "Yuzey", "Aciklama",
    ])
    for b in results.busbars:
        for h in b.holes:
            ws_hole.append([
                b.part_no, b.busbar_type, b.phase, h.hole_no,
                float(h.x_mm), float(h.y_mm),
                float(h.diameter_mm) if h.diameter_mm is not None else None,
                float(h.slot_width_mm)  if h.slot_width_mm  else None,
                float(h.slot_length_mm) if h.slot_length_mm else None,
                h.face or "", h.description or "",
            ])
    for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
        ws_hole.column_dimensions[col].width = 14

    # ── Bükümler ─────────────────────────────────────────────────────────────
    ws_bend = wb.create_sheet("Bukumler")
    _header_row(ws_bend, [
        "Parca No", "Tip", "Faz", "Bukum No",
        "Bastan Mesafe (mm)", "Aci (deg)", "Yon",
        "Ic Yari Cap (mm)", "Eksen", "Tur", "Bukum Payi (mm)", "Aciklama",
    ])
    for b in results.busbars:
        for bd in b.bends:
            ws_bend.append([
                b.part_no, b.busbar_type, b.phase, bd.bend_no,
                float(bd.distance_from_start_mm), float(bd.angle_deg), bd.direction,
                float(bd.inner_radius_mm),
                bd.bend_axis or "", bd.bend_type or "",
                float(bd.bend_allowance_mm) if bd.bend_allowance_mm is not None else None,
                bd.description or "",
            ])
    for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
        ws_bend.column_dimensions[col].width = 16

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


# ─── PDF ─────────────────────────────────────────────────────────────────────

def build_pdf(db: Session, project_id: int) -> bytes:
    results = get_results(db, project_id)
    out = BytesIO()
    c = canvas.Canvas(out, pagesize=A4)
    W, H = A4

    def new_page() -> float:
        c.showPage()
        return H - 40

    y = H - 40

    # Başlık
    c.setFont("Helvetica-Bold", 16)
    c.drawString(40, y, f"PanelForge — Proje {project_id}")
    y -= 30

    # Özet
    c.setFont("Helvetica-Bold", 12)
    c.drawString(40, y, "Ozet")
    y -= 18
    c.setFont("Helvetica", 10)
    s = results.summary
    for label, val in [
        ("Ana bakir sayisi:",       str(s.main_busbar_count)),
        ("Tali bakir sayisi:",      str(s.branch_busbar_count)),
        ("Toplam kesim boyu:",      f"{float(s.total_cut_length_mm):.1f} mm"),
        ("Toplam delik:",           str(s.total_hole_count)),
        ("Toplam bukum:",           str(s.total_bend_count)),
        ("Toplam agirlik:",         f"{float(s.total_weight_kg):.3f} kg"),
    ]:
        c.drawString(40, y, label)
        c.drawString(220, y, val)
        y -= 16
        if y < 60:
            y = new_page()

    # Uyarılar
    if results.warnings:
        y -= 10
        c.setFont("Helvetica-Bold", 11)
        c.drawString(40, y, "Geometri Uyarilari:")
        y -= 16
        c.setFont("Helvetica", 9)
        for w in results.warnings:
            c.setFillColorRGB(0.8, 0.1, 0.1)
            c.drawString(50, y, f"• {w}")
            c.setFillColorRGB(0, 0, 0)
            y -= 14
            if y < 60:
                y = new_page()

    y -= 20

    # Bakır parçaları
    c.setFont("Helvetica-Bold", 12)
    c.drawString(40, y, "Bakir Parcalari")
    y -= 20

    for busbar in results.busbars:
        if y < 100:
            y = new_page()
        c.setFont("Helvetica-Bold", 9)
        c.drawString(40, y,
            f"{busbar.part_no}  |  {busbar.busbar_type}  |  {busbar.phase}  "
            f"|  {busbar.width_mm}×{busbar.thickness_mm} mm  "
            f"|  {float(busbar.cut_length_mm):.1f} mm  "
            f"|  {busbar.material}"
        )
        y -= 14
        c.setFont("Helvetica", 8)
        # Delikler
        if busbar.holes:
            hole_strs = [f"H{h.hole_no}:x={float(h.x_mm):.1f}" for h in busbar.holes]
            c.drawString(55, y, "Delikler: " + "  ".join(hole_strs))
            y -= 13
        # Bükümler
        if busbar.bends:
            bend_strs = [
                f"B{bd.bend_no}:{float(bd.distance_from_start_mm):.1f}mm/{bd.direction}"
                for bd in busbar.bends
            ]
            c.drawString(55, y, "Bukumler: " + "  ".join(bend_strs))
            y -= 13
        y -= 5
        if y < 60:
            y = new_page()

    c.save()
    return out.getvalue()


# ─── DXF ─────────────────────────────────────────────────────────────────────

def build_dxf(db: Session, project_id: int) -> bytes:
    results = get_results(db, project_id)
    doc = ezdxf.new("R2010")
    msp = doc.modelspace()

    # Layer tanımları
    for layer_name, color in [
        ("BUSBAR_MAIN",   2),   # sarı
        ("BUSBAR_BRANCH", 4),   # cyan
        ("HOLE_MAIN",     1),   # kırmızı
        ("HOLE_BRANCH",   3),   # yeşil
    ]:
        doc.layers.add(layer_name, color=color)

    for b in results.busbars:
        layer_line = "BUSBAR_MAIN" if b.busbar_type == "main" else "BUSBAR_BRANCH"
        layer_hole = "HOLE_MAIN"   if b.busbar_type == "main" else "HOLE_BRANCH"

        # Segmentler — 3D çizgi
        for seg in b.segments:
            sx = float(seg.start_x_mm)
            sy = float(seg.start_y_mm)
            sz = float(getattr(seg, "start_z_mm", None) or 0)
            ex = float(seg.end_x_mm)
            ey = float(seg.end_y_mm)
            ez = float(getattr(seg, "end_z_mm", None) or 0)
            if sz == 0 and ez == 0:
                msp.add_line((sx, sy), (ex, ey), dxfattribs={"layer": layer_line})
            else:
                msp.add_line((sx, sy, sz), (ex, ey, ez), dxfattribs={"layer": layer_line})

        # Delikler — ana bara üzerindeki kavşak noktaları (segment başlangıcından ilerletilir)
        if b.busbar_type == "main" and b.segments:
            seg0 = b.segments[0]
            ox = float(seg0.start_x_mm)
            oy = float(seg0.start_y_mm)
            # Segment yönü (normalize)
            total_x = float(b.segments[-1].end_x_mm) - ox
            total_y = float(b.segments[-1].end_y_mm) - oy
            length = (total_x**2 + total_y**2) ** 0.5
            if length > 0:
                ux, uy = total_x / length, total_y / length
            else:
                ux, uy = 1.0, 0.0

            for h in b.holes:
                hx = float(h.x_mm)
                cx = ox + ux * hx
                cy = oy + uy * hx
                r = float(h.diameter_mm) / 2 if h.diameter_mm else 5.5
                msp.add_circle((cx, cy), radius=r, dxfattribs={"layer": layer_hole})

    output = StringIO()
    doc.write(output)
    return output.getvalue().encode("utf-8")
