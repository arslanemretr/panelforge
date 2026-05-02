from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy.orm import Session, selectinload

from app.db import models
from app.schemas.device import DeviceCreate, DeviceTerminalCreate
from app.schemas.device_import import (
    DeviceImportError,
    DeviceImportPreview,
    DeviceImportPreviewRow,
    DeviceImportResult,
)


DEVICE_HEADERS = [
    "device_code",
    "brand",
    "model",
    "device_type",
    "poles",
    "current_a",
    "width_mm",
    "height_mm",
    "depth_mm",
]

TERMINAL_HEADERS = [
    "device_code",
    "terminal_name",
    "phase",
    "x_mm",
    "y_mm",
    "z_mm",
    "terminal_face",
    "hole_diameter_mm",
    "slot_width_mm",
    "slot_length_mm",
]

VALID_PHASES = {"L1", "L2", "L3", "N", "PE"}
VALID_FACES = {"front", "back", "left", "right", "top", "bottom"}


def _clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return value


def _to_int(value: Any) -> int:
    cleaned = _clean(value)
    if cleaned is None:
        raise ValueError("zorunlu sayi alanı boş")
    return int(float(cleaned))


def _to_float(value: Any, *, required: bool = False, default: float | None = None) -> float | None:
    cleaned = _clean(value)
    if cleaned is None:
        if required:
            raise ValueError("zorunlu sayi alanı boş")
        return default
    return float(cleaned)


def _build_workbook(devices_rows: list[list[Any]], terminal_rows: list[list[Any]], notes: list[list[Any]] | None = None) -> bytes:
    workbook = Workbook()
    devices_sheet = workbook.active
    devices_sheet.title = "devices"
    devices_sheet.append(DEVICE_HEADERS)
    for row in devices_rows:
        devices_sheet.append(row)

    terminals_sheet = workbook.create_sheet("terminals")
    terminals_sheet.append(TERMINAL_HEADERS)
    for row in terminal_rows:
        terminals_sheet.append(row)

    if notes:
        notes_sheet = workbook.create_sheet("notes")
        for row in notes:
            notes_sheet.append(row)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_template_excel() -> bytes:
    devices_rows = [
        ["MCCB_1600_MAIN", "ABB", "Emax 2", "Ana Salter", 3, 1600, 260, 420, 180],
        ["MCCB_250_QF1", "ABB", "Tmax XT5", "Tali Salter", 3, 250, 140, 255, 120],
        ["MCCB_160_QF2", "Schneider", "Compact NSX", "Tali Salter", 3, 160, 135, 240, 110],
    ]
    terminal_rows = [
        ["MCCB_1600_MAIN", "L1_IN", "L1", 30, 20, 0, "front", 11, None, None],
        ["MCCB_1600_MAIN", "L2_IN", "L2", 90, 20, 0, "front", 11, None, None],
        ["MCCB_1600_MAIN", "L3_IN", "L3", 150, 20, 0, "front", 11, None, None],
        ["MCCB_1600_MAIN", "L1_OUT", "L1", 30, 390, 0, "front", 11, None, None],
        ["MCCB_1600_MAIN", "L2_OUT", "L2", 90, 390, 0, "front", 11, None, None],
        ["MCCB_1600_MAIN", "L3_OUT", "L3", 150, 390, 0, "front", 11, None, None],
        ["MCCB_250_QF1", "L1", "L1", 25, 20, 0, "front", 9, None, None],
        ["MCCB_250_QF1", "L2", "L2", 70, 20, 0, "front", 9, None, None],
        ["MCCB_250_QF1", "L3", "L3", 115, 20, 0, "front", 9, None, None],
        ["MCCB_160_QF2", "L1", "L1", 22, 20, 0, "front", 9, None, None],
        ["MCCB_160_QF2", "L2", "L2", 67, 20, 0, "front", 9, None, None],
        ["MCCB_160_QF2", "L3", "L3", 112, 20, 0, "front", 9, None, None],
    ]
    notes = [
        ["PanelForge toplu cihaz import ornek dosyasi"],
        [],
        ["devices sayfasi zorunlu alanlari", "device_code", "brand", "model", "device_type", "poles", "width_mm", "height_mm"],
        ["terminals sayfasi zorunlu alanlari", "device_code", "terminal_name", "phase", "x_mm", "y_mm"],
        ["phase kabul degerleri", "L1", "L2", "L3", "N", "PE"],
        ["terminal_face kabul degerleri", "front", "back", "left", "right", "top", "bottom"],
    ]
    return _build_workbook(devices_rows, terminal_rows, notes)


def build_export_excel(db: Session) -> bytes:
    devices = (
        db.query(models.Device)
        .options(selectinload(models.Device.terminals))
        .order_by(models.Device.id.asc())
        .all()
    )
    device_rows: list[list[Any]] = []
    terminal_rows: list[list[Any]] = []

    for device in devices:
        code = f"DEV_{device.id}"
        device_rows.append([
            code,
            device.brand,
            device.model,
            device.device_type,
            device.poles,
            float(device.current_a) if device.current_a is not None else None,
            float(device.width_mm),
            float(device.height_mm),
            float(device.depth_mm) if device.depth_mm is not None else None,
        ])
        for terminal in device.terminals:
            terminal_rows.append([
                code,
                terminal.terminal_name,
                terminal.phase,
                float(terminal.x_mm),
                float(terminal.y_mm),
                float(terminal.z_mm) if terminal.z_mm is not None else 0,
                terminal.terminal_face,
                float(terminal.hole_diameter_mm) if terminal.hole_diameter_mm is not None else None,
                float(terminal.slot_width_mm) if terminal.slot_width_mm is not None else None,
                float(terminal.slot_length_mm) if terminal.slot_length_mm is not None else None,
            ])

    return _build_workbook(device_rows, terminal_rows)


def _parse_workbook(file_bytes: bytes) -> DeviceImportPreview:
    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    except (InvalidFileException, KeyError, ValueError, TypeError) as exc:
        return DeviceImportPreview(
            can_import=False,
            device_count=0,
            terminal_count=0,
            errors=[DeviceImportError(sheet="workbook", row=1, message=f"Excel dosyasi okunamadi: {exc}")],
            devices=[],
        )
    errors: list[DeviceImportError] = []

    if "devices" not in workbook.sheetnames:
        return DeviceImportPreview(
            can_import=False,
            device_count=0,
            terminal_count=0,
            errors=[DeviceImportError(sheet="devices", row=1, message="devices sayfasi bulunamadi")],
            devices=[],
        )
    if "terminals" not in workbook.sheetnames:
        return DeviceImportPreview(
            can_import=False,
            device_count=0,
            terminal_count=0,
            errors=[DeviceImportError(sheet="terminals", row=1, message="terminals sayfasi bulunamadi")],
            devices=[],
        )

    devices_sheet = workbook["devices"]
    terminals_sheet = workbook["terminals"]

    device_headers = [str(cell).strip() if cell is not None else "" for cell in next(devices_sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    terminal_headers = [str(cell).strip() if cell is not None else "" for cell in next(terminals_sheet.iter_rows(min_row=1, max_row=1, values_only=True))]

    if device_headers != DEVICE_HEADERS:
        errors.append(DeviceImportError(sheet="devices", row=1, message="devices kolonlari beklenen sira ile eslesmiyor"))
    if terminal_headers != TERMINAL_HEADERS:
        errors.append(DeviceImportError(sheet="terminals", row=1, message="terminals kolonlari beklenen sira ile eslesmiyor"))

    devices_by_code: dict[str, dict[str, Any]] = {}
    terminals_by_code: dict[str, list[dict[str, Any]]] = {}

    for row_index, values in enumerate(devices_sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(_clean(v) is None for v in values):
            continue
        row = dict(zip(DEVICE_HEADERS, values))
        code = _clean(row["device_code"])
        if not isinstance(code, str):
            errors.append(DeviceImportError(sheet="devices", row=row_index, message="device_code zorunludur"))
            continue
        if code in devices_by_code:
            errors.append(DeviceImportError(sheet="devices", row=row_index, message=f"{code} birden fazla kez tanimli"))
            continue
        try:
            devices_by_code[code] = {
                "brand": str(_clean(row["brand"]) or ""),
                "model": str(_clean(row["model"]) or ""),
                "device_type": str(_clean(row["device_type"]) or ""),
                "poles": _to_int(row["poles"]),
                "current_a": _to_float(row["current_a"]),
                "width_mm": _to_float(row["width_mm"], required=True),
                "height_mm": _to_float(row["height_mm"], required=True),
                "depth_mm": _to_float(row["depth_mm"]),
            }
            if not devices_by_code[code]["brand"] or not devices_by_code[code]["model"] or not devices_by_code[code]["device_type"]:
                raise ValueError("brand, model ve device_type zorunludur")
        except ValueError as exc:
            errors.append(DeviceImportError(sheet="devices", row=row_index, message=f"{code}: {exc}"))

    terminal_count = 0
    for row_index, values in enumerate(terminals_sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(_clean(v) is None for v in values):
            continue
        row = dict(zip(TERMINAL_HEADERS, values))
        code = _clean(row["device_code"])
        if not isinstance(code, str):
            errors.append(DeviceImportError(sheet="terminals", row=row_index, message="device_code zorunludur"))
            continue
        phase = str(_clean(row["phase"]) or "").upper()
        face = _clean(row["terminal_face"])
        try:
            terminal = {
                "terminal_name": str(_clean(row["terminal_name"]) or ""),
                "phase": phase,
                "x_mm": _to_float(row["x_mm"], required=True),
                "y_mm": _to_float(row["y_mm"], required=True),
                "z_mm": _to_float(row["z_mm"], default=0.0),
                "terminal_face": str(face).lower() if isinstance(face, str) else None,
                "hole_diameter_mm": _to_float(row["hole_diameter_mm"]),
                "slot_width_mm": _to_float(row["slot_width_mm"]),
                "slot_length_mm": _to_float(row["slot_length_mm"]),
            }
            if code not in devices_by_code:
                raise ValueError(f"{code} devices sayfasinda tanimli degil")
            if not terminal["terminal_name"]:
                raise ValueError("terminal_name zorunludur")
            if terminal["phase"] not in VALID_PHASES:
                raise ValueError(f"phase gecersiz: {terminal['phase']}")
            if terminal["terminal_face"] and terminal["terminal_face"] not in VALID_FACES:
                raise ValueError(f"terminal_face gecersiz: {terminal['terminal_face']}")
            terminals_by_code.setdefault(code, []).append(terminal)
            terminal_count += 1
        except ValueError as exc:
            errors.append(DeviceImportError(sheet="terminals", row=row_index, message=str(exc)))

    preview_rows: list[DeviceImportPreviewRow] = []
    for code, device in devices_by_code.items():
        terminal_rows = terminals_by_code.get(code, [])
        if not terminal_rows:
            errors.append(DeviceImportError(sheet="terminals", row=1, message=f"{code} icin terminal tanimi yok"))
            continue
        preview_rows.append(
            DeviceImportPreviewRow(
                device_code=code,
                brand=device["brand"],
                model=device["model"],
                device_type=device["device_type"],
                poles=device["poles"],
                current_a=device["current_a"],
                width_mm=device["width_mm"] or 0,
                height_mm=device["height_mm"] or 0,
                depth_mm=device["depth_mm"],
                terminal_count=len(terminal_rows),
            )
        )

    return DeviceImportPreview(
        can_import=len(errors) == 0 and len(preview_rows) > 0,
        device_count=len(preview_rows),
        terminal_count=terminal_count,
        errors=errors,
        devices=preview_rows[:20],
    )


def preview_import(file_bytes: bytes) -> DeviceImportPreview:
    return _parse_workbook(file_bytes)


def import_devices_from_excel(db: Session, file_bytes: bytes) -> DeviceImportResult:
    preview = _parse_workbook(file_bytes)
    if not preview.can_import:
        raise ValueError("Dosya import icin uygun degil")

    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    devices_sheet = workbook["devices"]
    terminals_sheet = workbook["terminals"]

    terminals_by_code: dict[str, list[DeviceTerminalCreate]] = {}
    for values in terminals_sheet.iter_rows(min_row=2, values_only=True):
        if all(_clean(v) is None for v in values):
            continue
        row = dict(zip(TERMINAL_HEADERS, values))
        code = str(_clean(row["device_code"]))
        face = _clean(row["terminal_face"])
        terminals_by_code.setdefault(code, []).append(
            DeviceTerminalCreate(
                terminal_name=str(_clean(row["terminal_name"])),
                phase=str(_clean(row["phase"])).upper(),
                x_mm=_to_float(row["x_mm"], required=True) or 0,
                y_mm=_to_float(row["y_mm"], required=True) or 0,
                z_mm=_to_float(row["z_mm"], default=0.0) or 0,
                terminal_face=str(face).lower() if isinstance(face, str) else None,
                hole_diameter_mm=_to_float(row["hole_diameter_mm"]),
                slot_width_mm=_to_float(row["slot_width_mm"]),
                slot_length_mm=_to_float(row["slot_length_mm"]),
            )
        )

    created_devices = 0
    created_terminals = 0

    for values in devices_sheet.iter_rows(min_row=2, values_only=True):
        if all(_clean(v) is None for v in values):
            continue
        row = dict(zip(DEVICE_HEADERS, values))
        code = _clean(row["device_code"])
        if not code:
            continue
        payload = DeviceCreate(
            brand=str(_clean(row["brand"])),
            model=str(_clean(row["model"])),
            device_type=str(_clean(row["device_type"])),
            poles=_to_int(row["poles"]),
            current_a=_to_float(row["current_a"]),
            width_mm=_to_float(row["width_mm"], required=True) or 0,
            height_mm=_to_float(row["height_mm"], required=True) or 0,
            depth_mm=_to_float(row["depth_mm"]),
            terminals=terminals_by_code.get(str(code), []),
        )
        device = models.Device(**payload.model_dump(exclude={"terminals"}))
        db.add(device)
        db.flush()
        for terminal in payload.terminals:
            db.add(models.DeviceTerminal(device_id=device.id, **terminal.model_dump()))
            created_terminals += 1
        created_devices += 1

    db.commit()
    return DeviceImportResult(
        created_device_count=created_devices,
        created_terminal_count=created_terminals,
    )
